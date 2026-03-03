from __future__ import annotations

from io import BytesIO

from .asset_placeholders import collect_asset_occurrences
from .asset_resolver import asset_missing_error
from .image_utils import (
    DEFAULT_XLSX_MAX_HEIGHT_PX,
    DEFAULT_XLSX_MAX_WIDTH_PX,
    resolve_display_size,
)
from .office_asset_types import ResolvedOfficeAsset


def inject_xlsx_assets(
    *,
    payload: bytes,
    aliases_by_key: dict[str, list[str]],
    assets_by_key: dict[str, ResolvedOfficeAsset | None],
    fail_on_missing_asset: bool,
    error_by_key: dict[str, dict[str, object]] | None = None,
) -> tuple[bytes, list[dict[str, object]], dict[str, int]]:
    from openpyxl import load_workbook  # type: ignore
    from openpyxl.drawing.image import Image as XLImage  # type: ignore

    workbook = load_workbook(filename=BytesIO(payload))
    errors: list[dict[str, object]] = []
    occurrences: dict[str, int] = {}
    available_keys = set(aliases_by_key.keys())
    per_key_errors = error_by_key or {}

    for ws in workbook.worksheets:
        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                value = cell.value
                if not isinstance(value, str) or not value:
                    continue

                matches = collect_asset_occurrences(value, available_keys)
                if not matches:
                    continue

                updated = value
                for match in matches:
                    key = match.key
                    occurrences[key] = occurrences.get(key, 0) + 1
                    resolved = assets_by_key.get(key)
                    if resolved is None:
                        if fail_on_missing_asset:
                            errors.append(per_key_errors.get(key) or asset_missing_error(key))
                        continue

                    draw_w, draw_h = resolve_display_size(
                        source_width=resolved.width,
                        source_height=resolved.height,
                        request_width=match.width_px,
                        request_height=match.height_px,
                        max_width=DEFAULT_XLSX_MAX_WIDTH_PX,
                        max_height=DEFAULT_XLSX_MAX_HEIGHT_PX,
                    )
                    image = XLImage(BytesIO(resolved.payload))
                    image.width = draw_w
                    image.height = draw_h
                    ws.add_image(image, cell.coordinate)

                    updated = updated.replace(match.raw, "")

                if fail_on_missing_asset and any(assets_by_key.get(match.key) is None for match in matches):
                    continue
                cell.value = updated.strip() or None

    out = BytesIO()
    workbook.save(out)
    return out.getvalue(), errors, occurrences
